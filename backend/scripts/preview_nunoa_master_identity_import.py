"""Preview import plan for Ñuñoa Master civil identity data.

This script is intentionally read-only: it reads the workbook and, when
DATABASE_URL is available, compares against the target database without writing.
It produces CSV/JSON artifacts that can be reviewed before any real import.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover - preview still works without DB checks.
    psycopg = None
    dict_row = None


REQUIRED_COLUMNS = {
    "Rut",
    "Nombres",
    "1er apellido",
    "2do apellido",
    "apellido, nombre",
    "Fecha de nacimiento",
    "CORREO",
    "genero",
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
RUT_RE = re.compile(r"^[0-9]{7,8}[0-9K]$")
GENDER_MAP = {"M": "male", "F": "female"}
NON_BLOCKING_ISSUES = {"missing_rut"}


@dataclass
class WorkbookMember:
    row_number: int
    rut_raw: str
    rut_normalized: str | None
    first_name: str
    first_surname: str
    second_surname: str
    last_name: str
    competition_name: str
    date_of_birth: str | None
    email: str | None
    gender: str | None
    issues: list[str] = field(default_factory=list)


@dataclass
class DbState:
    club_id: int | None = None
    club_name: str | None = None
    existing_person_by_rut: dict[str, int] = field(default_factory=dict)
    existing_memberships: set[tuple[int, int]] = field(default_factory=set)
    existing_contacts: set[tuple[int, str, str]] = field(default_factory=set)


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_rut(value: Any) -> str | None:
    raw = clean(value).upper().replace(".", "").replace("-", "")
    return raw or None


def is_valid_rut(rut: str | None) -> bool:
    if not rut or not RUT_RE.match(rut):
        return False
    body, expected_dv = rut[:-1], rut[-1]
    total = 0
    multiplier = 2
    for digit in reversed(body):
        total += int(digit) * multiplier
        multiplier = 2 if multiplier == 7 else multiplier + 1
    value = 11 - (total % 11)
    calculated_dv = "0" if value == 11 else "K" if value == 10 else str(value)
    return calculated_dv == expected_dv


def normalize_email(value: Any) -> str | None:
    email = clean(value).lower()
    return email or None


def parse_date(value: Any) -> str | None:
    if value is None or clean(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Fecha no reconocida: {text!r}")


def load_members(workbook_path: Path) -> list[WorkbookMember]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_COLUMNS.difference(headers)
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {sorted(missing)}")

    members: list[WorkbookMember] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(clean(value) for value in values):
            continue
        row = dict(zip(headers, values))
        rut = normalize_rut(row.get("Rut"))
        email = normalize_email(row.get("CORREO"))
        gender_raw = clean(row.get("genero")).upper()
        gender = GENDER_MAP.get(gender_raw)
        first_name = clean(row.get("Nombres"))
        first_surname = clean(row.get("1er apellido"))
        second_surname = clean(row.get("2do apellido"))
        last_name = " ".join(part for part in (first_surname, second_surname) if part)
        date_of_birth = parse_date(row.get("Fecha de nacimiento"))

        member = WorkbookMember(
            row_number=row_number,
            rut_raw=clean(row.get("Rut")),
            rut_normalized=rut,
            first_name=first_name,
            first_surname=first_surname,
            second_surname=second_surname,
            last_name=last_name,
            competition_name=clean(row.get("apellido, nombre")),
            date_of_birth=date_of_birth,
            email=email,
            gender=gender,
        )

        if not rut:
            member.issues.append("missing_rut")
        elif not is_valid_rut(rut):
            member.issues.append("invalid_rut")
        if email and not EMAIL_RE.match(email):
            member.issues.append("invalid_email")
        if gender is None:
            member.issues.append(f"unknown_gender:{gender_raw}")
        if not first_name or not first_surname:
            member.issues.append("incomplete_name")

        # The workbook's competition-name column is expected to be first surname
        # plus first given name; flag drift so aliases are reviewed before import.
        if first_name and first_surname:
            expected_competition_name = f"{first_surname}, {first_name.split()[0]}".upper()
            if member.competition_name.upper() != expected_competition_name:
                member.issues.append("competition_name_mismatch")

        members.append(member)
    return members


DbConnectInfo = str | dict[str, str | int] | None


def load_db_state(database_url: DbConnectInfo, club_name: str) -> DbState:
    state = DbState()
    if not database_url:
        return state
    if psycopg is None or dict_row is None:
        raise RuntimeError("psycopg no está instalado; no se puede comparar contra la BD.")

    if isinstance(database_url, dict):
        conn_ctx = psycopg.connect(**database_url, row_factory=dict_row)
    else:
        conn_ctx = psycopg.connect(database_url, row_factory=dict_row)

    with conn_ctx as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, name
                FROM core.club
                WHERE LOWER(TRIM(name)) = LOWER(TRIM(%s))
                   OR LOWER(TRIM(short_name)) = LOWER(TRIM(%s))
                ORDER BY id
                LIMIT 1
                """,
                (club_name, club_name),
            )
            club = cur.fetchone()
            if club:
                state.club_id = int(club["id"])
                state.club_name = str(club["name"])

            cur.execute(
                """
                SELECT id, rut_normalized
                FROM identity.person
                WHERE rut_normalized IS NOT NULL
                """
            )
            state.existing_person_by_rut = {
                str(row["rut_normalized"]): int(row["id"]) for row in cur.fetchall()
            }

            cur.execute(
                """
                SELECT person_id, contact_type, LOWER(TRIM(contact_value)) AS contact_value
                FROM identity.contact_point
                """
            )
            state.existing_contacts = {
                (int(row["person_id"]), str(row["contact_type"]), str(row["contact_value"]))
                for row in cur.fetchall()
            }

            if state.club_id is not None:
                cur.execute(
                    """
                    SELECT club_id, person_id
                    FROM club_ops.membership
                    WHERE club_id = %s
                    """,
                    (state.club_id,),
                )
                state.existing_memberships = {
                    (int(row["club_id"]), int(row["person_id"])) for row in cur.fetchall()
                }
    return state


def db_connect_kwargs(database_url: str | None) -> str | dict[str, str | int] | None:
    if database_url:
        return database_url
    host = os.getenv("DB_HOST")
    name = os.getenv("DB_NAME")
    user = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    if not (host and name and user and password):
        return None
    return {
        "host": host,
        "port": int(os.getenv("DB_PORT", "5432")),
        "dbname": name,
        "user": user,
        "password": password,
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = sorted({key for row in rows for key in row})
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_preview(members: list[WorkbookMember], db: DbState) -> dict[str, Any]:
    people: list[dict[str, Any]] = []
    contacts: list[dict[str, Any]] = []
    memberships: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []

    for member in members:
        blocking_issues = [
            issue for issue in member.issues if issue not in NON_BLOCKING_ISSUES
        ]
        existing_person_id = (
            db.existing_person_by_rut.get(member.rut_normalized)
            if member.rut_normalized
            else None
        )
        person_action = "review" if blocking_issues else "insert_person"
        if not blocking_issues and not member.rut_normalized:
            person_action = "insert_person_missing_rut"
        if existing_person_id and not blocking_issues:
            person_action = "update_or_keep_person"

        row_base = asdict(member)
        row_base["issues"] = "|".join(member.issues)
        row_base["blocking_issues"] = "|".join(blocking_issues)
        row_base["existing_person_id"] = existing_person_id
        row_base["person_action"] = person_action

        people.append(row_base)
        if blocking_issues:
            review.append(row_base)
            continue

        planned_person_id = existing_person_id
        if member.email:
            contact_exists = (
                bool(planned_person_id)
                and (planned_person_id, "email", member.email) in db.existing_contacts
            )
            contacts.append(
                {
                    "row_number": member.row_number,
                    "person_id": planned_person_id,
                    "rut_normalized": member.rut_normalized,
                    "contact_type": "email",
                    "contact_value": member.email,
                    "action": "keep_contact" if contact_exists else "insert_contact",
                }
            )

        membership_exists = (
            bool(db.club_id and planned_person_id)
            and (db.club_id, planned_person_id) in db.existing_memberships
        )
        memberships.append(
            {
                "row_number": member.row_number,
                "person_id": planned_person_id,
                "rut_normalized": member.rut_normalized,
                "club_id": db.club_id,
                "club_name": db.club_name,
                "status": "active",
                "action": "keep_membership" if membership_exists else "insert_membership",
            }
        )

    return {
        "people": people,
        "contacts": contacts,
        "memberships": memberships,
        "review": review,
        "summary": {
            "rows": len(members),
            "auto_ready_people": sum(1 for row in people if row["person_action"] != "review"),
            "requires_review": len(review),
            "missing_rut_non_blocking": sum(
                1 for row in people if "missing_rut" in row["issues"].split("|")
            ),
            "planned_contacts": len(contacts),
            "planned_memberships": len(memberships),
            "db_club_id": db.club_id,
            "db_club_name": db.club_name,
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Preview Ñuñoa Master identity import.")
    parser.add_argument(
        "--workbook",
        default="backend/data/staging/Ñuñoa Master 2026.xlsx",
        help="Path to the Ñuñoa Master workbook.",
    )
    parser.add_argument(
        "--club-name",
        default="Ñuñoa Master",
        help="Club name/short_name to resolve in core.club when DATABASE_URL exists.",
    )
    parser.add_argument(
        "--output-dir",
        default="backend/data/staging/nunoa_master_identity_preview",
        help="Directory for preview CSV/JSON outputs.",
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help="Optional database URL. Defaults to DATABASE_URL from backend/.env/environment.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    backend_dir = Path(__file__).resolve().parents[1]
    load_dotenv(backend_dir / ".env")
    workbook_path = Path(args.workbook)
    output_dir = Path(args.output_dir)
    database_url = db_connect_kwargs(args.database_url or os.getenv("DATABASE_URL"))

    members = load_members(workbook_path)
    db = load_db_state(database_url, args.club_name)
    preview = build_preview(members, db)

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "people_preview.csv", preview["people"])
    write_csv(output_dir / "contacts_preview.csv", preview["contacts"])
    write_csv(output_dir / "memberships_preview.csv", preview["memberships"])
    write_csv(output_dir / "requires_review.csv", preview["review"])
    (output_dir / "summary.json").write_text(
        json.dumps(
            {
                "state": "preview_generated",
                "workbook": str(workbook_path),
                "output_dir": str(output_dir),
                **preview["summary"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(json.dumps(preview["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
